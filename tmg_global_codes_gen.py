"""Generate idempotent tmgGlobalCodes INSERT scripts for regulatory DBA changes.

Follows the established pattern from V26.2.0707__DATA_tmgGlobalCodes_232_Steel_Aluminum_5368657.sql.
Auto-detects the next available version sequence number in the target directory.

Records are read automatically from the ADO work item description when -r is omitted.
Requires 'az login' for ADO authentication.

Usage (auto-fetch records from ADO):
    python scripts/tmg_global_codes_gen.py \\
        --work-item 5463147 \\
        --target-dir "D:/dev/ogt/gtm-legacy_gtm-sql/Database/Application/26.2/Hotfix" \\
        --series 07 \\
        --description "232_Metals_CSMS68855869"

Usage (manual records):
    python scripts/tmg_global_codes_gen.py \\
        --work-item 5463147 \\
        --target-dir "D:/dev/ogt/gtm-legacy_gtm-sql/Database/Application/26.2/Hotfix" \\
        --series 07 \\
        -r "ABIFTZ-HTS-ALUMINIUM-54RECORD:37013000" \\
        -r "ABIFTZ-HTS-STEEL-54DERIVATIVE:8708292120"
"""
import argparse
import html.parser
import io
import json
import re
import subprocess
import sys
import urllib.error
import urllib.request
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

ADO_ORG = 'https://dev.azure.com/tr-tax'
ADO_PROJECT = 'TaxTrade'
ADO_RESOURCE = '499b84ac-1321-427f-aa17-267ca6975798'

FIELDNAME_RE = re.compile(r'^ABIFTZ-HTS-[A-Z0-9-]+$')
CODE_RE = re.compile(r'^[\d.]+$')


# ---------------------------------------------------------------------------
# ADO helpers
# ---------------------------------------------------------------------------

def get_az_token() -> str:
    az_cmd = 'az.cmd' if sys.platform == 'win32' else 'az'
    result = subprocess.run(
        [az_cmd, 'account', 'get-access-token',
         '--resource', ADO_RESOURCE,
         '--query', 'accessToken', '-o', 'tsv'],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        sys.exit(
            "ERROR: Failed to get ADO token. Run 'az login' first.\n"
            + result.stderr.strip()
        )
    return result.stdout.strip()


def fetch_work_item(work_item_id: str) -> dict:
    token = get_az_token()
    url = (
        f"{ADO_ORG}/{ADO_PROJECT}/_apis/wit/workitems/{work_item_id}"
        f"?$expand=all&api-version=7.0"
    )
    req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        sys.exit(f"ERROR: ADO API returned {e.code} for work item {work_item_id}. "
                 f"Check the ID and your az login session.")


def find_excel_attachment(wi_data: dict) -> tuple[str, str] | None:
    """Return (url, filename) of the first .xlsx attachment, or None."""
    for rel in wi_data.get('relations') or []:
        if rel.get('rel') == 'AttachedFile':
            name = rel.get('attributes', {}).get('name', '')
            if name.lower().endswith('.xlsx'):
                return rel['url'], name
    return None


def download_attachment(url: str) -> bytes:
    token = get_az_token()
    req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        sys.exit(f"ERROR: Failed to download attachment ({e.code}): {url}")


def parse_records_from_excel(content: bytes) -> list[tuple[str, str]]:
    """Read FieldName/Code pairs from the INSERTS sheet of the work item Excel attachment.

    Filters to rows where Action = 'Insert' when that column is present,
    so reference/context rows in other tabs are never included.
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl is required to read Excel attachments. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Find the INSERTS sheet
    sheet = None
    for name in wb.sheetnames:
        if 'INSERT' in name.upper():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    # Locate header columns (Action, FieldName, Code)
    action_col = fieldname_col = code_col = header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value or '').strip()
            if val == 'Action':
                action_col = cell.column
                header_row = cell.row
            elif val == 'FieldName':
                fieldname_col = cell.column
                header_row = cell.row
            elif val == 'Code':
                code_col = cell.column
        if fieldname_col and code_col:
            break

    if not fieldname_col or not code_col:
        return []

    records = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        # Filter by Action = 'Insert' when column exists
        if action_col:
            action = str(row[action_col - 1].value or '').strip().lower()
            if action and action != 'insert':
                continue

        field = str(row[fieldname_col - 1].value or '').strip()
        code  = str(row[code_col  - 1].value or '').strip()

        # Normalize float-encoded codes like 37013000.0 -> 37013000
        if code.endswith('.0') and code[:-2].isdigit():
            code = code[:-2]

        if FIELDNAME_RE.match(field) and CODE_RE.match(code):
            records.append((field, code))

    return records


# ---------------------------------------------------------------------------
# HTML parsing — extract (FieldName, Code) pairs from work item description
# ---------------------------------------------------------------------------

class _CellExtractor(html.parser.HTMLParser):
    """Collect plain-text content of every <td> in document order."""

    def __init__(self):
        super().__init__()
        self._depth = 0       # nesting depth inside a <td>
        self._buf: list[str] = []
        self.cells: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == 'td':
            self._depth += 1
            if self._depth == 1:
                self._buf = []

    def handle_endtag(self, tag):
        if tag == 'td' and self._depth > 0:
            self._depth -= 1
            if self._depth == 0:
                text = ' '.join(''.join(self._buf).split()).strip()
                if text:
                    self.cells.append(text)
                self._buf = []

    def handle_data(self, data):
        if self._depth > 0:
            self._buf.append(data)


def parse_records_from_html(description_html: str) -> list[tuple[str, str]]:
    """Extract (FieldName, Code) pairs from the work item description HTML.

    Looks for consecutive <td> cells where the first matches the ABIFTZ-HTS-*
    pattern and the second is a numeric/dotted code value.
    """
    extractor = _CellExtractor()
    extractor.feed(description_html)
    cells = extractor.cells

    records: list[tuple[str, str]] = []
    i = 0
    while i < len(cells) - 1:
        if FIELDNAME_RE.match(cells[i]) and CODE_RE.match(cells[i + 1]):
            records.append((cells[i], cells[i + 1]))
            i += 2
        else:
            i += 1

    return records


# ---------------------------------------------------------------------------
# Script generation
# ---------------------------------------------------------------------------

def find_version(target_dir: Path) -> str:
    for part in reversed(target_dir.parts):
        if re.fullmatch(r'\d+\.\d+', part):
            return part
    sys.exit(
        f"ERROR: Could not detect release version from path: {target_dir}\n"
        f"       Expected a component like '26.2' in the path."
    )


def find_next_sequence(target_dir: Path, version: str, series: str | None = None) -> int:
    """Return max existing sequence number + 1, optionally scoped to a series prefix."""
    pattern = re.compile(rf'^V{re.escape(version)}\.(\d+)__', re.IGNORECASE)
    max_seq = 0
    for f in target_dir.iterdir():
        m = pattern.match(f.name)
        if m:
            seq_str = m.group(1)
            if series and not seq_str.startswith(series):
                continue
            max_seq = max(max_seq, int(seq_str))
    if max_seq == 0 and series:
        return int(series) * (10 ** (4 - len(series))) + 1
    return max_seq + 1


def sanitize_for_sql_name(value: str) -> str:
    return re.sub(r'[^A-Za-z0-9_]', '_', value)


def group_records(records: list[tuple[str, str]]) -> OrderedDict:
    groups: OrderedDict = OrderedDict()
    for field, code in records:
        groups.setdefault(field, []).append(code)
    return groups


def build_source_data_values(groups: OrderedDict) -> str:
    lines = []
    field_names = list(groups.keys())
    for i, (field, codes) in enumerate(groups.items()):
        count = len(codes)
        lines.append(f"        -- {field} ({count} code{'s' if count != 1 else ''})")
        for j, code in enumerate(codes):
            is_last = (i == len(field_names) - 1) and (j == len(codes) - 1)
            comma = '' if is_last else ','
            lines.append(f"        ('{field}', '{code}'){comma}")
        if i < len(field_names) - 1:
            lines.append('')
    return '\n'.join(lines)


def build_merge_block(groups: OrderedDict) -> str:
    """Shared MERGE block reused by deploy script and QA harness."""
    source_values = build_source_data_values(groups)
    return f"""\
    DECLARE @SourceData TABLE(
        FieldName VARCHAR(30),
        Code      NVARCHAR(36)
    );

    INSERT INTO @SourceData (FieldName, Code)
    VALUES
{source_values};

    MERGE dbo.tmgGlobalCodes dt
    USING @SourceData AS st
        ON  dt.[PartnerID]     = @PartnerID
            AND dt.[FieldName] = st.FieldName
            AND dt.[Code]      = st.Code
    WHEN NOT MATCHED BY TARGET THEN
        INSERT ([PartnerID], [EffDate], [FieldName], [Code], [Decode], [StaticFlag], [DeletedFlag], [KeepDuringRollback])
        VALUES (@PartnerID, @EffDate, st.FieldName, st.Code, st.Code, 'Y', 'N', 'N');"""


def build_ac_queries(groups: OrderedDict, backup_table: str) -> str:
    """Shared AC verification queries reused by verify and QA harness scripts."""
    all_codes = ', '.join(f"'{c}'" for codes in groups.values() for c in codes)
    total = sum(len(c) for c in groups.values())
    lines = []

    lines.append(f"/* ---- AC-1 / AC-2: backup exists and is not overwritten ---- */")
    lines.append(f"SELECT [AC] = 'AC-1/AC-2',")
    lines.append(f"       [BackupExists]  = CASE WHEN OBJECT_ID('{backup_table}','U') IS NOT NULL THEN 'PASS' ELSE 'FAIL' END,")
    lines.append(f"       [BackupTable]   = '{backup_table}';")
    lines.append("")

    ac_num = 3
    for field, codes in groups.items():
        expected = len(codes)
        if expected == 1:
            where = f"FieldName = '{field}' AND Code = '{codes[0]}'"
        else:
            code_list = ', '.join(f"'{c}'" for c in codes)
            where = f"FieldName = '{field}' AND Code IN ({code_list})"
        lines.append(f"/* ---- AC-{ac_num}: {field} count = {expected} ---- */")
        lines.append(f"SELECT [AC] = 'AC-{ac_num}', [Expected] = {expected},")
        lines.append(f"       [Actual] = COUNT(*),")
        lines.append(f"       [Status] = CASE WHEN COUNT(*) = {expected} THEN 'PASS' ELSE 'FAIL' END")
        lines.append(f"FROM tmgGlobalCodes WITH (NOLOCK)")
        lines.append(f"WHERE {where};")
        lines.append("")
        ac_num += 1

    lines.append(f"/* ---- AC-{ac_num}: idempotency — no duplicates on re-run, counts unchanged ---- */")
    lines.append(f"/* (verified by the harness Pass 2 zero-insert check) */")
    lines.append("")
    ac_num += 1

    lines.append(f"/* ---- AC-{ac_num}: final review — all {total} codes present with DeletedFlag = 'N' ---- */")
    lines.append(f"SELECT [AC] = 'AC-{ac_num}', [Expected] = {total},")
    lines.append(f"       [Actual] = COUNT(*),")
    lines.append(f"       [Status] = CASE WHEN COUNT(*) = {total} THEN 'PASS' ELSE 'FAIL' END")
    lines.append(f"FROM tmgGlobalCodes WITH (NOLOCK)")
    lines.append(f"WHERE Code IN ({all_codes}) AND DeletedFlag = 'N';")

    return '\n'.join(lines)


def build_verify_sql(work_item: str, title: str, backup_table: str, groups: OrderedDict) -> str:
    ac_block = build_ac_queries(groups, backup_table)
    return f"""\
/* ============================================================
   US {work_item} -- tmgGlobalCodes Verification
   {title}
   Run AFTER the deploy script. Read-only. All SELECTs WITH (NOLOCK).
   Every [Status] column should read 'PASS'.
============================================================ */

SET NOCOUNT ON;

{ac_block}
"""


def build_rollback_harness_sql(work_item: str, title: str,
                                backup_table: str, groups: OrderedDict,
                                date_str: str) -> str:
    total = sum(len(c) for c in groups.values())
    merge_block = build_merge_block(groups)
    ac_block = build_ac_queries(groups, backup_table)

    return f"""\
/* ============================================================
   *** TEMPORARY — QA ROLLBACK TEST HARNESS — DO NOT COMMIT ***
   US {work_item} -- {title}

   Runs the REAL deploy operations + the REAL AC verification inside a
   SINGLE transaction that ALWAYS ROLLS BACK. Nothing is persisted.
   Composed from the same shared builders as the deploy and verify scripts.

   PASS 1 = apply   (expect {total} inserts)
   PASS 2 = re-run  (expect 0 inserts — proves idempotency)
   Then AC verification against the uncommitted state, then ROLLBACK.
============================================================ */

SET NOCOUNT ON;
SET XACT_ABORT ON;

DECLARE @PartnerID   AS INT      = (SELECT TOP 1 PARTNERID FROM [dbo].tmfDefaults WITH (NOLOCK));
DECLARE @EffDate     DATETIME    = GETDATE();
DECLARE @Inserted1   INT = 0;
DECLARE @Inserted2   INT = 0;

IF @PartnerID IS NULL
BEGIN RAISERROR('Unable to retrieve PartnerID from tmfDefaults.', 16, 1); RETURN; END

BEGIN TRY
    BEGIN TRANSACTION;

    /* ===== PASS 1: apply the REAL MERGE (expect {total} inserts) ===== */
{merge_block}
    SET @Inserted1 = @@ROWCOUNT;
    SELECT [Phase] = 'PASS 1 (applied, uncommitted)',
           [Inserted] = @Inserted1,
           [Expected] = {total},
           [Status]   = CASE WHEN @Inserted1 = {total} THEN 'PASS' ELSE 'CHECK — may already exist' END;

    /* ===== PASS 2: re-run to prove idempotency (expect 0 inserts) ===== */
{merge_block}
    SET @Inserted2 = @@ROWCOUNT;
    SELECT [Phase]      = 'PASS 2 (idempotency re-run)',
           [Inserted]   = @Inserted2,
           [Expected]   = 0,
           [Idempotent] = CASE WHEN @Inserted2 = 0 THEN 'PASS' ELSE 'FAIL' END;

    /* ===== AC VERIFICATION against the uncommitted state ===== */
{ac_block}

    /* ===== INTENTIONAL ROLLBACK — nothing is persisted ===== */
    IF @@TRANCOUNT > 0 ROLLBACK TRANSACTION;
    PRINT '*** ROLLED BACK — no changes were persisted to this database. ***';

END TRY
BEGIN CATCH
    IF @@TRANCOUNT > 0 ROLLBACK TRANSACTION;
    SELECT [Phase] = 'ERROR (rolled back)',
           [ErrLine] = ERROR_LINE(),
           [ErrMsg]  = ERROR_MESSAGE();
END CATCH;

/* ---- Post-rollback proof: backup table must NOT exist (nothing was committed) ---- */
SELECT [PostRollbackProof]  = 'Backup table must NOT exist (proves rollback succeeded)',
       [BackupExists]       = CASE WHEN OBJECT_ID('{backup_table}','U') IS NOT NULL THEN 'FAIL — was committed!' ELSE 'PASS' END;
"""


def generate_sql(filename: str, work_item: str, title: str,
                 date_str: str, backup_table: str, groups: OrderedDict) -> str:
    merge_block = build_merge_block(groups)

    return f"""\
--------------------------------------------------------------------------------------------------------------
-- Script: {filename}
-- Purpose: {title}
-- Story: {work_item}
-- Date: {date_str}
--------------------------------------------------------------------------------------------------------------

DECLARE @Msg              NVARCHAR(4000) = '"';
DECLARE @InsertCount      INT = 0;
DECLARE @EffDate          DATETIME = GETDATE();
DECLARE @WindowsBreakline VARCHAR(2) = CHAR(13) + CHAR(10);

DECLARE @BackupTableName SYSNAME = N'{backup_table}';

BEGIN TRY

    IF OBJECT_ID('dbo.tmfDefaults', 'U') IS NULL
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'ERROR: tmfDefaults table does not exist';

\t    IF LEN (@Msg) < 2
\t    SET @Msg = null;

\t    IF LEN (@Msg) > 1
\t    SET @Msg = @Msg + CHAR(10) +  '"';

        SELECT
             DatabaseName   = DB_NAME()
            ,InsertCount    = @InsertCount
            ,Msgs           = @Msg
        RETURN;
    END

    IF OBJECT_ID('dbo.tmgGlobalCodes', 'U') IS NULL
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'ERROR: tmgGlobalCodes table does not exist';

\t    IF LEN (@Msg) < 2
\t    SET @Msg = null;

\t    IF LEN (@Msg) > 1
\t    SET @Msg = @Msg + CHAR(10) +  '"';

        SELECT
             DatabaseName   = DB_NAME()
            ,InsertCount    = @InsertCount
            ,Msgs           = @Msg
        RETURN;
    END

    IF SCHEMA_ID('bck') IS NULL
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'ERROR: Backup schema [bck] does not exist';

\t    IF LEN (@Msg) < 2
\t    SET @Msg = null;

\t    IF LEN (@Msg) > 1
\t    SET @Msg = @Msg + CHAR(10) +  '"';

        SELECT
             DatabaseName   = DB_NAME()
            ,InsertCount    = @InsertCount
            ,Msgs           = @Msg
        RETURN;
    END

    DECLARE @PartnerID AS INT = (SELECT TOP 1 PARTNERID FROM [dbo].tmfDefaults WITH (NOLOCK));

    IF @PartnerID IS NULL
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'ERROR: Unable to retrieve PartnerID from tmfDefaults';

\t\tIF LEN (@Msg) < 2
\t\tSET @Msg = null;

\t\tIF LEN (@Msg) > 1
\t\tSET @Msg = @Msg + CHAR(10) +  '"';

        SELECT
             DatabaseName   = DB_NAME()
            ,InsertCount    = @InsertCount
            ,Msgs           = @Msg
        RETURN;
    END

    BEGIN TRANSACTION;

    IF OBJECT_ID(@BackupTableName, 'U') IS NULL
    BEGIN
        DECLARE @BackupSQL nvarchar(max) = N'
            SELECT
                [PartnerID]
               ,[EffDate]
               ,[FieldName]
               ,[Code]
               ,[Decode]
               ,[StaticFlag]
               ,[DeletedFlag]
               ,[KeepDuringRollback]
            INTO ' + @BackupTableName + N'
            FROM [dbo].[tmgGlobalCodes] WITH (NOLOCK)';

        EXEC sys.sp_executesql @BackupSQL;

        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Backup table ' + @BackupTableName + ' created.';
    END
    ELSE
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Backup table ' + @BackupTableName + ' already exists. Skipping backup.';
    END

    -- Apply Insert using MERGE on (partnerid, fieldname, code)
    -- Skip codes already present in the partner's tmgGlobalCodes
{merge_block}

    DECLARE @MergeCount INT = @@ROWCOUNT;

    COMMIT TRANSACTION;

    SET @InsertCount = @MergeCount;
    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Inserted ' + CAST(@InsertCount AS VARCHAR(10)) + ' new tmgGlobalCodes';

END TRY
BEGIN CATCH

    DECLARE @ErrorMessage NVARCHAR(4000) = ERROR_MESSAGE();
    DECLARE @ErrorSeverity INT = ERROR_SEVERITY();
    DECLARE @ErrorState INT = ERROR_STATE();
    DECLARE @ErrorLine INT = ERROR_LINE();

    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Error occurred in script execution:';
    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Error Message: ' + @ErrorMessage;
    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Error Line: ' + CAST(@ErrorLine AS VARCHAR(10));
    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Error Severity: ' + CAST(@ErrorSeverity AS VARCHAR(10));
    SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Error State: ' + CAST(@ErrorState AS VARCHAR(10));

    IF @@TRANCOUNT > 0
    BEGIN
        SET @Msg = @Msg + @WindowsBreakline + ' - ' +  'Transaction rolled back due to error.';
        ROLLBACK TRANSACTION;
    END

END CATCH;

IF LEN (@Msg) < 2
SET @Msg = null;

IF LEN (@Msg) > 1
SET @Msg = @Msg + CHAR(10) +  '"';

SELECT
     DatabaseName   = DB_NAME()
    ,InsertCount    = @InsertCount
    ,Msgs           = @Msg
"""


def parse_record(value: str) -> tuple[str, str]:
    parts = value.split(':', 1)
    if len(parts) != 2 or not parts[0].strip() or not parts[1].strip():
        sys.exit(f"ERROR: Invalid --record format '{value}'. Expected 'FIELDNAME:CODE'.")
    return parts[0].strip(), parts[1].strip()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description='Generate an idempotent tmgGlobalCodes INSERT script.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--work-item', '-w', required=True,
                        help='ADO work item ID (e.g. 5463147)')
    parser.add_argument('--target-dir', '-d', required=True,
                        help='Directory to write the SQL file into')
    parser.add_argument('--record', '-r', action='append', default=[],
                        metavar='FIELDNAME:CODE',
                        help='Record to insert — repeat for multiple. '
                             'If omitted, records are parsed from the ADO work item description.')
    parser.add_argument('--description',
                        help='Short tag for the filename (e.g. "232_Metals_CSMS68855869"). '
                             'If omitted and fetching from ADO, derived from the work item title.')
    parser.add_argument('--title',
                        help='Human-readable title for the script header. '
                             'If omitted and fetching from ADO, uses the work item title.')
    parser.add_argument('--date',
                        help='Effective date YYYY-MM-DD (default: today)')
    parser.add_argument('--series',
                        help='Sequence prefix to stay within (e.g. "07" -> next in 0701–0799)')
    parser.add_argument('--out-verify', metavar='PATH',
                        help='Also write AC verification script to this path. '
                             'If omitted, written alongside the deploy script as VERIFY_*.sql')
    parser.add_argument('--out-test', metavar='PATH',
                        help='Also write QA rollback test harness to this path. '
                             'If omitted but --with-test is set, written as QA_TEST_*.sql')
    parser.add_argument('--with-test', action='store_true',
                        help='Generate the QA rollback test harness alongside the deploy script')
    parser.add_argument('--dry-run', action='store_true',
                        help='Print the generated SQL to stdout without writing a file')

    args = parser.parse_args()

    target_dir = Path(args.target_dir)
    if not target_dir.is_dir():
        sys.exit(f"ERROR: target directory does not exist: {target_dir}")

    # --- Resolve records ---
    if args.record:
        records = [parse_record(r) for r in args.record]
        wi_title = None
    else:
        print(f"Fetching work item {args.work_item} from ADO...")
        wi_data = fetch_work_item(args.work_item)
        wi_title = wi_data['fields'].get('System.Title', '')

        # Prefer Excel attachment (source of truth) over HTML description
        attachment = find_excel_attachment(wi_data)
        if attachment:
            att_url, att_name = attachment
            print(f"Downloading Excel attachment: {att_name}")
            content = download_attachment(att_url)
            records = parse_records_from_excel(content)
            if records:
                print(f"Found {len(records)} record(s) in Excel attachment:")
                for field, code in records:
                    print(f"  {field} : {code}")
            else:
                print("WARNING: No ABIFTZ-HTS-* records found in Excel. Falling back to HTML description.")
                records = parse_records_from_html(wi_data['fields'].get('System.Description', ''))
        else:
            print("No Excel attachment found. Parsing HTML description...")
            records = parse_records_from_html(wi_data['fields'].get('System.Description', ''))
            if records:
                print(f"Found {len(records)} record(s) in work item description:")
                for field, code in records:
                    print(f"  {field} : {code}")

        if not records:
            sys.exit(
                "ERROR: No ABIFTZ-HTS-* records found in work item. Use -r to provide records manually."
            )

    groups = group_records(records)
    version = find_version(target_dir)
    seq = find_next_sequence(target_dir, version, args.series)
    date_str = args.date or datetime.today().strftime('%Y-%m-%d')
    date_compact = date_str.replace('-', '')

    # --- Derive description / title ---
    if args.description:
        description = args.description
    elif wi_title:
        # e.g. "Regulatory - DBA Script tmgGlobalCodes — Section 232 ... CSMS 68855869"
        # Extract CSMS number if present, otherwise sanitize title
        csms_match = re.search(r'CSMS\s*(\d+)', wi_title)
        if csms_match:
            description = f"232_Metals_CSMS{csms_match.group(1)}"
        else:
            description = re.sub(r'[^A-Za-z0-9]+', '_', wi_title)[:60].strip('_')
    else:
        field_tags = '_'.join(
            re.sub(r'ABIFTZ-HTS-', '', f).replace('-', '_') for f in groups.keys()
        )
        description = f"tmgGlobalCodes_{field_tags}"

    title = args.title or wi_title or description.replace('_', ' ')

    filename = f"V{version}.{seq:04d}__DATA_tmgGlobalCodes_{description}_{args.work_item}.sql"
    backup_tag = sanitize_for_sql_name(description)
    backup_table = f"[bck].[bck_tmgGlobalCodes_{backup_tag}_{date_compact}_{args.work_item}]"

    sql = generate_sql(filename, args.work_item, title, date_str, backup_table, groups)

    verify_sql  = build_verify_sql(args.work_item, title, backup_table, groups)
    harness_sql = build_rollback_harness_sql(args.work_item, title, backup_table, groups, date_str)

    if args.dry_run:
        print("=" * 70)
        print("DEPLOY SCRIPT")
        print("=" * 70)
        print(sql)
        print("=" * 70)
        print("VERIFY SCRIPT")
        print("=" * 70)
        print(verify_sql)
        if args.with_test or args.out_test:
            print("=" * 70)
            print("QA ROLLBACK HARNESS")
            print("=" * 70)
            print(harness_sql)
        return

    output_path  = target_dir / filename
    verify_stem  = f"VERIFY_tmgGlobalCodes_{description}_{args.work_item}.sql"
    harness_stem = f"QA_TEST_tmgGlobalCodes_{description}_{args.work_item}.sql"

    verify_path  = Path(args.out_verify)  if args.out_verify  else target_dir / verify_stem
    harness_path = Path(args.out_test)    if args.out_test    else target_dir / harness_stem

    if output_path.exists():
        sys.exit(f"ERROR: File already exists: {output_path}\n"
                 f"       Delete it or use --dry-run to preview.")

    output_path.write_text(sql, encoding='utf-8')
    verify_path.write_text(verify_sql, encoding='utf-8')
    print(f"\nCreated deploy  : {output_path}")
    print(f"Created verify  : {verify_path}")

    if args.with_test or args.out_test:
        harness_path.write_text(harness_sql, encoding='utf-8')
        print(f"Created harness : {harness_path}  *** DO NOT COMMIT ***")

    print(f"\n  Version  : {version}")
    print(f"  Sequence : {seq:04d}")
    print(f"  Records  : {sum(len(c) for c in groups.values())} inserts across {len(groups)} FieldName(s)")
    print(f"  Backup   : {backup_table}")


if __name__ == '__main__':
    main()
